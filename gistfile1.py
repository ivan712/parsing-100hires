from myutils import *
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
from time import sleep
from openpyxl import load_workbook
import requests

from random import randint
import json


current_position = list()
current_position.append([
    'Talent',
    'Recruiter',
    'Recruitment',
    'Recruiting',
    'Sourcing',
    'Sourcer'])

current_position.append([
    'people',
    'HR',
    'human'
])

current_position.append([
    'Operations',
    'Assistant',
    'Office'
])

current_position.append([
    'COO',
    'CEO',
    'Founder',
    'Co-Founder'
])

RUCAPTCHA_KEY = "################"


add_input_script = """
var inp = document.createElement('input');
inp.type = 'submit';
inp.value = 'send';
inp.id = 'send_token';
document.getElementById('captcha-form').appendChild(inp);
"""


def solve_recaptcha(browser):
    browser.execute_script(add_input_script)
    send_token_input = browser.find_element_by_id('send_token')
    text_area_for_token = browser.find_element_by_id('g-recaptcha-response')
    browser.execute_script("document.getElementById('g-recaptcha-response').style.display = 'inline';")
    cookies = browser.get_cookies()
    cookies_to_send = str()
    for cookie in cookies:
        # print(cookie)
        for key in cookie.keys():
            cookies_to_send += f"{key}:{cookie[key]};"
    html = browser.page_source
    bs_obj = BeautifulSoup(html, 'html5lib')
    recaptca_tag = bs_obj.find('div', {'class': 'g-recaptcha', 'id': 'recaptcha'})

    data_sitekey = recaptca_tag['data-sitekey']
    data_s = recaptca_tag['data-s']
    data_callback = recaptca_tag['data-callback']
    page_url = browser.current_url
    req_str = f"https://rucaptcha.com/in.php?" \
              f"key={RUCAPTCHA_KEY}&" \
              f"method=userrecaptcha&" \
              f"googlekey={data_sitekey}&" \
              f"data-s={data_s}&" \
              f"cookies={cookies_to_send}&" \
              f"pageurl={page_url}&" \
              f"json=1&" \
              f"debug_dumps=1"
    # if we want to use proxy, we should use this request url
    '''
    req_str = f"https://rucaptcha.com/in.php?" \
              f"key={RUCAPTCHA_KEY}&" \
              f"method=userrecaptcha&" \
              f"googlekey={data_sitekey}&" \
              f"data-s={data_s}&" \
              f"proxy={AUTH}@{PROXY}&" \
              f"proxytype=HTTPS&" \
              f"pageurl={page_url}&" \
              f"json=1&" \
              f"debug_dumps=1"
    '''
    req_ans = requests.get(req_str)
    response = req_ans.text
    response = json.loads(response)
    if response['status'] == 1:
        id = response['request']
        req_res = f"https://rucaptcha.com/res.php?" \
                  f"key={RUCAPTCHA_KEY}&" \
                  f"action=get&" \
                  f"id={id}&" \
                  f"json=1"
        print("Our request is processing")
        print(f"id = {id}")
        while True:
            sleep(20)
            res = requests.get(req_res).text
            res = json.loads(res)
            if res['status'] == 1:
                print("Captcha is solved successfully")
                token = res['request']
                add_cookies = res['cookies']
                for key in add_cookies.keys():
                    if add_cookies[key] == 'True':
                        add_cookies[key] = True
                        continue
                    if add_cookies[key] == 'False':
                        add_cookies[key] = False
                        continue
                    if add_cookies[key].isdigit():
                        add_cookies[key] = int(add_cookies[key])
                text_area_for_token.send_keys(token)
                send_token_input.click()
                return True
            if res['request'] == 'ERROR_CAPTCHA_UNSOLVABLE':
                browser.refresh()
                solve_recaptcha(browser)
                break
            print(f"{res['request']} -- Waiting")


def check_current_position(title):
    for cur_pos in current_position:
        for pos in cur_pos:
            if pos in title:
                return True
    return False


def get_google_search_res(browser, query):
    search_box = browser.find_element_by_name('q')
    search_box.send_keys(Keys.CONTROL + 'a')
    search_box.send_keys(Keys.DELETE)
    search_box.send_keys(query)
    sleep(2)
    search_box.submit()
    sleep(2)
    scroll = randint(350, 450)
    browser.execute_script(f'window.scrollTo(0,{scroll});')
    sleep(randint(2, 7))
    html = browser.page_source
    bs_obj = BeautifulSoup(html, 'html5lib')
    res_tag = bs_obj.find('div', {'id': 'res'})
    while True:
        if not res_tag:
            solve_recaptcha(browser)
            html = browser.page_source
            bs_obj = BeautifulSoup(html, 'html5lib')
            res_tag = bs_obj.find('div', {'id': 'res'})
        else:
            res_tags = res_tag.find_all('div', {'class': 'g'})
            break
    res = list()
    for res_tag in res_tags:
        h3 = res_tag.h3.get_text()
        splitted_h3 = h3.split(' - ')
        if len(splitted_h3) == 1:
            another_split = h3.split(' – ')[0]
            if len(another_split) == 1:
                continue
            name = another_split[0]
            job_title = another_split[1]
        else:
            name = splitted_h3[0]
            job_title = splitted_h3[1]
        # если у нас нет текущих слов в job title
        # то пропускаем
        if not check_current_position(job_title):
            continue
        link = ''
        if 'href' in res_tag.a.attrs:
            link = res_tag.a['href']
        res.append({'name': name, 'title': job_title, 'link': link})

    return res


def get_comp_title(browser, linkedin_url):
    while True:
        try:
            load_linkedin_page(browser, linkedin_url)
        except:
            print(f"Couldn't load {linkedin_url}")
            input("Press Enter to continue")
        else:
            break
    html = browser.page_source
    bs_obj = BeautifulSoup(html, 'html5lib')
    title = bs_obj.find('h1', {'class': 'org-top-card-summary__title t-24 t-black t-bold truncate'}).get_text().strip()
    return title


def comp_titles():
    from json import loads
    wb = load_workbook('linkedin_for_titles.xlsx')
    ws = wb.active
    with open('accs.txt', 'r') as f:
        proxies = loads(f.read())
    with open('parsed.txt', 'r') as f:
        parsed = f.read()
    proxy = proxies["1"]
    browser = init_driver(gui=True, profile=proxy['profile'], proxy=proxy['proxy'], auth=proxy['proxy-auth'])
    for i in range(2, 5000):
        linkedin_url = ws[f'C{i}'].value
        if not linkedin_url or linkedin_url in parsed:
            continue
        print(f'{i} - {linkedin_url}')
        try:
            title = get_comp_title(browser, linkedin_url)
        except:
            fprint('Something went wrong')
            with open('parsed.txt', 'a') as f:
                f.write(linkedin_url + '\n')
        else:
            with open('parsed.txt', 'a') as f:
                f.write(linkedin_url + '\n')
            with open('res.csv', 'a') as f:
                f.write(linkedin_url + ';' + title + '\n')


def find_names():
    with open('parsed_name.txt', 'r') as f:
        parsed = f.read().split('\n')
    with open('res.csv', 'r') as f:
        lines = f.read().split('\n')
    # browser = init_driver(gui=True, profile="/home/toozox/tmp/chromium/google-robert-lynch", proxy=PROXY, auth=AUTH)
    # browser = init_driver(gui=True, profile="/home/toozox/tmp/chromium/google-robert-lynch")
    browser = init_driver(gui=True, profile='/tmp/aeu')
    browser.get('https://www.google.com')
    counter = 1
    for line in lines:
        comp_title = line.split(';')[1]
        linkdin_url = line.split(';')[0]
        if linkdin_url in parsed:
            counter += 1
            continue
        print(f"{counter} -- {comp_title}")
        cur_pos_str = '" OR "'.join(current_position[0])
        cur_pos_str = f'"{cur_pos_str}"'
        query = f'(intitle:"{comp_title}") (intitle:{cur_pos_str})  site:linkedin.com/in/'
        res = get_google_search_res(browser, query)
        for r in res:
            with open('res_names.csv', 'a') as f:
                f.write(f"{r['name']};{r['title']};{r['link']}\n")
        with open('parsed_name.txt', 'a') as f:
            f.write(linkdin_url + '\n')
        counter += 1
        sleep(2)


if __name__ == "__main__":
    # comp_titles()
    find_names()
